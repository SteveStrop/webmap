import datetime
import re
import configparser
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook


class EstateAgentJob:
    def __init__(self,
                 client='',
                 agent='',
                 href='',
                 ref_num='',
                 floorplan=True,
                 photos=True,
                 contact='',
                 phone_primary='',
                 phone_secondary='',
                 phone_other='',
                 email='',
                 address='',
                 postcode='',
                 appointment=''
                 ):
        self.client = client
        self.agent = Agent(agent)
        self.href = href
        self.ref_num = ref_num
        self.floorplan = floorplan
        self.photos = photos
        self.contact = contact
        self.phone_primary = phone_primary
        self.phone_secondary = phone_secondary
        self.phone_other = phone_other
        self.email = email
        self.appointment = Appointment(appointment, address, postcode)
        self.map_color = self.set_map_color()

    def set_map_color(self):  # TODO put webmap.ini colors directly in here
        if not self.appointment.date:
            return 'Tba'
        elif (self.appointment.date - datetime.datetime.now().date()).days > 14:
            return 'Later'
        elif (self.appointment.date - datetime.datetime.now().date()).days > 7:
            return self.appointment.date.strftime('%A')[:3] + '2'
        else:
            return self.appointment.date.strftime('%A')[:3] + '1'


class KAJob(EstateAgentJob):
    def __init__(self,
                 agent='',
                 notes='',
                 btn_fast_upload='',  # CSS class or id
                 btn_confirm_floorplan='',  # CSS class or id
                 btn_confirm_photos='',  # CSS class or id
                 btn_back='',  # CSS class or id
                 btn_save_apt='',  # CSS class or id
                 btn_change_apt='',  # CSS class or id
                 btn_save_no_apt='',  # CSS class or id
                 client='Key Agent',
                 href='',
                 ref_num='',
                 floorplan=True,
                 photos=True,
                 contact='',
                 phone_primary='',
                 phone_secondary='',
                 phone_other='',
                 email='',
                 address='',
                 postcode='',
                 appointment=''
                 ):
        super().__init__(client, agent, href, ref_num, floorplan, photos, contact, phone_primary, phone_secondary,
                         phone_other, email, address, postcode, appointment
                         )
        self.notes = notes
        self.btn_fast_upload = btn_fast_upload
        self.btn_confirm_floorplan = btn_confirm_floorplan
        self.btn_confirm_photos = btn_confirm_photos
        self.btn_back = btn_back
        self.btn_save_apt = btn_save_apt
        self.btn_change_apt = btn_change_apt
        self.btn_save_no_apt = btn_save_no_apt


class HSJob(EstateAgentJob):
    def __init__(self,
                 # add HSS job specifics here
                 client='House Simple',
                 href='',
                 ref_num='',
                 floorplan=True,
                 photos=True,
                 contact='',
                 phone_primary='',
                 phone_secondary='',
                 phone_other='',
                 email='',
                 address='',
                 postcode='',
                 appointment=''
                 ):
        # convert HS appointment format to match expected input of super class
        appointment = self.__convert_appointment(appointment)
        if not postcode:
            postcode = self.__get_postcode(address)

        super().__init__(client, href, ref_num, floorplan, photos, contact, phone_primary, phone_secondary, phone_other,
                         email,
                         address, postcode, appointment
                         )

    @staticmethod
    def __convert_appointment(appointment):
        """ cleans up input appointment string and returns a string suitable for __get__date_time method of super class
        """
        appointment = appointment.replace('/', '-')
        appointment = appointment.replace('@', '')
        return appointment

    @staticmethod
    def __get_postcode(address):
        """ extracts first occurrence of a valid UK postcode from a string"""
        try:
            return re.findall(r'[A-Z]{1,2}[\dR][\dA-Z]? [\d][A-Z]{2}', address)[0].strip()
        except:
            return ''


class Appointment:
    def __init__(self, appointment, address, postcode):
        self.address = address.replace(',', '')
        self.postcode = postcode
        self.__get__date_time(appointment)

    def __get__date_time(self, appointment):
        """ takes an appointment string 'dd-mm-yyyy hh:mm'.
        Creates a datetime date object from the first part (split on white space)
        and a time object from the second part"""
        # strip out leading and trailing whitespace and split on internal space
        appointment = appointment.strip().split()
        if appointment:
            # split first part on '-' and reverse order for input to datetime.date
            self.date = datetime.date(*map(int, reversed(appointment[0].split('-'))))
            self.display_date = self.date.strftime('%a %b %d')
            # split second part on ':' and reverse order for input to datetime.time
            self.time = datetime.time(*map(int, appointment[1].split(':')))
            self.display_time = self.time.strftime('%H:%M')
        else:
            self.date = ''
            self.time = ''
            self.display_date = ''
            self.display_time = ''


class Agent:
    def __init__(self, agent, tel=''):
        self.name = agent
        self.tel = tel if tel else ''


class Scraper:

    def __init__(self, agent, mode=None):
        self.mode = mode
        config = configparser.ConfigParser()  # TODO can't have config hard coded here
        config.read('webmap.ini')
        self.agent = config[agent]
        self.chrome_driver_path = config['PATH']['DRIVER_PATH']
        self.wb_path = config['PATH']['EXCEL_PATH']
        if self.mode == 'headless':
            self.chrome_options = Options()
            self.chrome_options.add_argument("--headless")
        else:
            self.chrome_options = None

        # connect selenium to the browser in headless mode
        print('\nOpening browser...')  # , 'NO BROWSER FOR EXCEL TESTING!!! DONT FORGET\n' * 8)
        self.browser = webdriver.Chrome(self.chrome_driver_path, chrome_options=self.chrome_options)

    def logon(self):
        """ logon and navigate to agent['landing_pg']
        """
        username = self.agent['USERNAME']
        password = self.agent['PASSWORD']
        login_pg = self.agent['LOGIN_PG']
        username_field = self.agent['USERNAME_FIELD']
        landing_pg = self.agent['LANDING_PG']
        password_field = self.agent['PASSWORD_FIELD']
        login_btn = self.agent['LOGIN_BTN']
        # Navigate to the application home page
        self.browser.get(login_pg)
        # get input fields
        username_field = self.browser.find_element_by_name(username_field)
        password_field = self.browser.find_element_by_name(password_field)
        # enter table
        username_field.send_keys(username)
        password_field.send_keys(password)
        # get the Login button & click it
        self.browser.find_element_by_name(login_btn).click()
        # navigate to the landing page
        self.browser.get(landing_pg)

    def html_table_read(self):
        """
        Reads table from html table assuming first row are column headings
        :return: list of dicts, one for each row k,v = column heading : table
        """

        # create a dictionary of table headings
        table_name = self.agent['JOBS_TABLE']
        headings = self.browser.find_elements_by_css_selector(table_name + ' th')
        headings_dict = {i: (heading.text.strip() or 'Col:{}'.format(i)) for i, heading in enumerate(headings)}
        rows = self.browser.find_elements_by_css_selector(table_name + ' tr')
        # remove first row containing headings
        rows.pop(0)

        line_data = [{headings_dict[i]: data.text for i, data in enumerate(row.find_elements_by_css_selector('td'))} for
                     row
                     in rows]
        return line_data

    def get_outlook_jobs(self):

        def _agent(i):
            agent = self.ws.cell(i + 5, 1).value
            regex = r'(?:AGENT.*?)([A-Z].*?)(?:([A-Z]{1,2}[\dR][\dA-Z]? [\d][A-Z]{2}))'
            match = re.search(regex, agent)
            if match:
                return match.group(1).strip()
            return ''

        def _postcode(address):
            regex = r'[A-Z]{1,2}[\dR][\dA-Z]? [\d][A-Z]{2}'
            match = re.findall(regex, address)
            if match:
                return match[0].strip()
            return ''

        def _address(i):
            address = self.ws.cell(i + 2, 1).value.strip()
            postcode = _postcode(address)
            address = address.replace(postcode, '').strip().strip(',')
            return address, postcode

        def _appointment(i):
            appointment = self.ws.cell(i + 3, 1).value
            if appointment:
                appointment = appointment[3:].replace('/', '-').strip()
            return appointment

        def _client(i):
            return self.ws.cell(i + 4, 1).value.strip('*').strip()

        def _ref_num(i):
            ref_num = self.ws.cell(i + 5, 1).value
            regex = r'^\d{10}'
            match = re.search(regex, ref_num)
            if match:
                return match.group(0).strip()
            return ''

        # try:
        wb = load_workbook(filename=self.wb_path, read_only=True)
        # get the first sheet in the workbook
        self.ws = wb[wb.sheetnames[0]]
        #
        row_count = float(self.ws.calculate_dimension().split('A').pop())  # get the last element in the list
        if row_count % 5:
            raise Exception  # TODO implement this properly 'Non integer number of outlook appointments. Re run
            # agent map'#
        else:
            jobs_list = []
            for i in range(0, int(row_count), 5):
                client = _client(i)
                address, postcode = _address(i)
                agent = _agent(i)
                appointment = _appointment(i)
                ref_num = _ref_num(i)

                job = EstateAgentJob(
                    client=client,
                    address=address,
                    agent=agent,
                    postcode=postcode,
                    appointment=appointment,
                    ref_num=ref_num
                )

                jobs_list.append(job)
        # # except:
        # pass
        # finally:
        wb.close()
        return jobs_list

    def get_ka_jobs(self):  # TODO move these into Scraper
        #  only jobs on key agent website without an appointment date are unbooked.
        #  All others are in booked_jobs.txt file created by MS Outlook
        # get a headless selenium browser object and agent specific constants from config file
        def _make_ka_job(row, ka):
            """ creates a KAJob class with address, agent, postcode and appointment details"""
            return KAJob(
                address=row[ka['ADDRESS']],  # from [KA] section of config file
                agent=row[ka['AGENT']],
                postcode=row[ka['POSTCODE']],
                appointment=row[ka['APPOINTMENT']],
                ref_num=row[ka['REF_NUM']]
            )

        print('\nLogging on...')
        self.logon()
        print('\nReading table...')
        table = self.html_table_read()
        jobs_list = [_make_ka_job(row, self.agent) for row in
                     table]  # if row[scraper.agent['APPOINTMENT']]]  # TODO negate the if condition to select non
        # booked
        #  jobs only
        self.browser.close()
        return jobs_list

    @staticmethod
    def remove_duplicates(main_list, sub_list):
        print('Removing duplicates...')
        for main in main_list:
            for i, sub in enumerate(sub_list):
                if main.ref_num == sub.ref_num:
                    del sub_list[i]
        return main_list, sub_list

    @staticmethod
    def display_jobs(jobs):
        """ displays sub set of scrapped data for debugging"""
        for job in jobs:
            print(f'{job.client:15}{job.appointment.address:80}{job.appointment.postcode:10}\
                  {job.agent.name:30}{job.appointment.display_date} {job.appointment.display_time:7}{job.map_color:9}')
